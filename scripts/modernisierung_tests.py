import itertools
import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Konstanten aus fachlichen Vorgaben
EV_KWH_PER_YEAR = 2550  # 17 kWh/100km * 15000 km
EV_CO2_MIX = 0.35
COMBUSTION_CO2 = 2415  # kg/a
COMBUSTION_FUEL_COST = 1940  # EUR/a
EV_CO2 = 893  # kg/a
EV_CO2_SAVING = COMBUSTION_CO2 - EV_CO2
CLIMATE_EXTRA = 450
HEATPUMP_EXTRA = 5500
MIN_GRID_IMPORT = 200
MIN_FEEDIN_SHARE = 0.30

@dataclass
class TestResult:
    inputs: Dict
    scenario: str
    outputs: Dict
    issues: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)

def load_data() -> Dict:
    data_path = Path(__file__).resolve().parent.parent / "data" / "data.json"
    with data_path.open(encoding="utf-8") as f:
        return json.load(f)

def roof_limit_kwp(roof_area: float) -> float:
    return max(0, int(roof_area // 7))

def pv_house_limit(house_type: str) -> float:
    limits = {"reihenhaus": 12, "doppelhaus": 15, "einfamilienhaus": 20}
    return limits.get(house_type, 15)

def recommend_pv_kwp(total_load: float, roof_area: float, house_type: str) -> float:
    base = max(6, total_load / 900)
    pv_raw = round(base, 1)
    max_roof = roof_limit_kwp(roof_area)
    pv_final = min(pv_raw, max_roof, pv_house_limit(house_type))
    return max(0, pv_final)

def recommend_battery_kwh(total_load: float, pv_kwp: float, pv_yield: float) -> float:
    daily_load = total_load / 365
    batt = max(5, min(12, daily_load * 0.8))
    # Speicher nie > 2 Tageserträge
    daily_pv = (pv_kwp * pv_yield) / 365
    return min(batt, daily_pv * 2)

def estimate_energy_balance(pv_kwp: float, battery_kwh: float, annual_load: float, pv_yield: float,
                            has_ev: bool, ev_load: float) -> Tuple[float, float, float]:
    pv_generation = pv_kwp * pv_yield
    direct_share = 0.32 if battery_kwh > 0 else 0.27
    direct_self = min(annual_load * direct_share, pv_generation * 0.9)
    pv_surplus = max(0, pv_generation - direct_self)

    battery_roundtrip = 0.83
    battery_daily = battery_kwh * 0.7
    annual_batt_input = min(pv_surplus, battery_daily * 365)
    battery_output = annual_batt_input * battery_roundtrip

    potential_self = direct_self + battery_output
    max_autarky = 0.75 if battery_kwh > 0 else 0.4
    autarky = min(max_autarky, potential_self / annual_load if annual_load else 0)
    self_use = autarky * annual_load

    feed_in = max(0, pv_generation - self_use)
    # Mindest-Einspeisung
    min_feed = pv_generation * MIN_FEEDIN_SHARE
    if feed_in < min_feed:
        feed_in = min_feed
        self_use = max(0, pv_generation - feed_in)
    grid = max(0, annual_load - self_use)

    ev_from_batt = 0
    if has_ev and battery_kwh > 0:
        ev_from_batt = min(ev_load * 0.7, battery_output * 0.4)

    return grid, feed_in, autarky * 100, ev_from_batt

def build_inputs_matrix() -> List[Dict]:
    house_types = ["reihenhaus", "doppelhaus", "einfamilienhaus"]
    areas = [100, 150, 200]
    people = [1, 3, 5]
    floor_heating = [False, True]
    insulation = ["schlecht", "normal", "gut"]
    roofs = [30, 50, 80]
    climate = [False, True]
    wallbox = [False, True]
    combos = itertools.product(house_types, areas, people, floor_heating, insulation, roofs, climate, wallbox)
    matrix = []
    for h, a, p, f, ins, roof, ac, wb in combos:
        matrix.append(
            {
                "houseType": h,
                "area": a,
                "people": p,
                "floorHeating": f,
                "insulation": ins,
                "roofArea": roof,
                "climate": ac,
                "wallbox": wb,
            }
        )
    return matrix

def calc_consumption_blocks(base_data: Dict, inp: Dict) -> Dict:
    house_key = "freistehend" if inp["houseType"] == "einfamilienhaus" else inp["houseType"]
    heating_per_sqm = base_data["consumption"]["heating_per_sqm"][house_key][inp["insulation"]]
    household = inp["people"] * base_data["consumption"]["per_person"]
    heating = inp["area"] * heating_per_sqm
    climate = CLIMATE_EXTRA if inp["climate"] else 0
    ev = EV_KWH_PER_YEAR if inp["wallbox"] else 0
    hp = HEATPUMP_EXTRA
    return {"household": household, "heating": heating, "climate": climate, "ev": ev, "heatpump": hp}

def scenario_calculations(base_data: Dict, inp: Dict) -> List[TestResult]:
    blocks = calc_consumption_blocks(base_data, inp)
    el_price = base_data["prices"]["electricity_eur_per_kwh"]
    gas_price = base_data["prices"]["gas_eur_per_kwh"]
    feed_in_tariff = base_data["prices"]["feed_in_eur_per_kwh"]
    pv_yield = base_data["pv"]["yield_per_kwp"]

    baseline_electric = blocks["household"]
    baseline_gas = blocks["heating"]
    baseline_cost = baseline_electric * el_price + baseline_gas * gas_price + (COMBUSTION_FUEL_COST if inp["wallbox"] else 0)
    co2_today = baseline_electric * base_data["co2"]["electricity_factor"] + baseline_gas * base_data["co2"]["gas_factor"]
    if inp["wallbox"]:
        co2_today += COMBUSTION_CO2

    scenarios = [
        ("Nur Photovoltaik", False, False),
        ("PV + Speicher", True, False),
        ("PV + Speicher + Wärmepumpe", True, True),
    ]

    results: List[TestResult] = []
    for label, use_batt, use_hp in scenarios:
        household_block = blocks["household"]
        climate_block = blocks["climate"]
        ev_block = blocks["ev"]
        hp_block = blocks["heatpump"] if use_hp else 0
        annual_consumption = household_block + climate_block + ev_block + hp_block
        heating_demand = 0 if use_hp else blocks["heating"]

        pv_kwp = recommend_pv_kwp(annual_consumption, inp["roofArea"], inp["houseType"])
        battery_kwh = recommend_battery_kwh(annual_consumption, pv_kwp, pv_yield) if use_batt else 0

        grid_import, feed_in, autarky_pct, ev_from_batt = estimate_energy_balance(
            pv_kwp,
            battery_kwh,
            annual_consumption,
            pv_yield,
            inp["wallbox"],
            ev_block,
        )

        pv_generation = pv_kwp * pv_yield
        pv_cost = pv_kwp * base_data["pv"]["cost_per_kwp"]
        battery_cost = battery_kwh * base_data["battery"]["cost_per_kwh"]
        hp_power = blocks["heatpump"] / base_data["heatpump"]["full_load_hours"]
        hp_cost = hp_power * base_data["heatpump"]["cost_per_kw"]
        total_cost = pv_cost + battery_cost + (hp_cost if use_hp else 0)

        post_el_cost = grid_import * el_price - feed_in * feed_in_tariff
        post_cost = post_el_cost + heating_demand * gas_price
        savings = baseline_cost - post_cost

        break_even = None
        if savings > 0:
            break_even = total_cost / savings

        # CO2 nachher: EV-Teil mit Strommix, Rest mit Stromfaktor
        ev_grid_share = min(ev_block, grid_import) if inp["wallbox"] else 0
        other_grid = grid_import - ev_grid_share
        co2_after = other_grid * base_data["co2"]["electricity_factor"] + ev_grid_share * EV_CO2_MIX + heating_demand * base_data["co2"]["gas_factor"]
        co2_saving = co2_today - co2_after

        outputs = {
            "pv_kwp": round(pv_kwp, 2),
            "battery_kwh": round(battery_kwh, 2),
            "grid_import": round(grid_import, 0),
            "feed_in": round(feed_in, 0),
            "autarky_pct": round(autarky_pct, 1),
            "pv_generation": round(pv_generation, 0),
            "co2_today": round(co2_today, 1),
            "co2_after": round(co2_after, 1),
            "co2_saving": round(co2_saving, 1),
            "break_even_years": round(break_even, 1) if break_even else None,
            "annual_cost_post": round(post_cost, 0),
            "total_cost": round(total_cost, 0),
            "household_block": household_block,
            "climate_block": climate_block,
            "ev_block": ev_block,
            "heatpump_block": hp_block,
            "heating_demand": blocks["heating"],
            "ev_from_batt": round(ev_from_batt, 0),
        }

        res = TestResult(inputs=inp, scenario=label, outputs=outputs)
        validate_rules(res, inp, use_batt, use_hp, pv_yield)
        results.append(res)

    return results

def validate_rules(res: TestResult, inp: Dict, use_batt: bool, use_hp: bool, pv_yield: float) -> None:
    o = res.outputs
    roof_max = roof_limit_kwp(inp["roofArea"])
    if o["pv_kwp"] > roof_max + 1e-6:
        res.issues.append(f"PV-Dimensionierung überschreitet Dachlimit ({o['pv_kwp']} kWp > {roof_max})")
    daily_pv = (o["pv_kwp"] * pv_yield) / 365 if pv_yield else 0
    if use_batt and o["battery_kwh"] > daily_pv * 2 + 1e-6:
        res.issues.append("Speicher größer als 2 Tageserträge")

    if o["grid_import"] < MIN_GRID_IMPORT:
        res.warnings.append("Netzbezug zu niedrig (<200 kWh/a)")

    expected_climate = CLIMATE_EXTRA if inp["climate"] else 0
    expected_ev = EV_KWH_PER_YEAR if inp["wallbox"] else 0
    expected_hp = HEATPUMP_EXTRA if use_hp else 0
    if o["climate_block"] != expected_climate:
        res.issues.append("Klima-Verbrauch nicht sauber getrennt")
    if o["ev_block"] != expected_ev:
        res.issues.append("EV-Verbrauch nicht sauber getrennt")
    if o["heatpump_block"] != expected_hp:
        res.issues.append("WP-Verbrauch nicht sauber getrennt")

    if inp["wallbox"]:
        if o["co2_saving"] <= 0:
            res.issues.append("CO2-Bilanz verschlechtert sich trotz E-Auto – bitte Rechenkern und Annahmen prüfen.")
        elif o["co2_saving"] < EV_CO2_SAVING - 400:
            res.warnings.append("CO2-Einsparung durch EV deutlich geringer als erwartet – Annahmen zu Strommix, PV-Anteil oder Fahrleistung prüfen.")

    autarky = o["autarky_pct"]
    if not use_batt and not use_hp and not (12 <= autarky <= 50):
        res.warnings.append("Autarkie außerhalb 12–50 % (Nur PV)")
    if use_batt and not use_hp and not (35 <= autarky <= 85):
        res.warnings.append("Autarkie außerhalb 35–85 % (PV+Speicher)")
    if use_batt and use_hp and not (45 <= autarky <= 90):
        res.warnings.append("Autarkie außerhalb 45–90 % (PV+Speicher+WP)")
    if autarky > 95 or autarky < 3:
        res.issues.append("Autarkie außerhalb physikalischer Grenzen (>95 % oder <3 %)")

    be = o["break_even_years"]
    if be is None or be <= 0:
        res.issues.append("Break-even nicht berechenbar oder negative Einsparung.")
    elif be > 40:
        res.warnings.append("Break-even sehr lang (>40 Jahre) – wirtschaftlich schwach.")

def to_dataframe(results: List[TestResult]) -> pd.DataFrame:
    rows = []
    for r in results:
        row = {
            **r.inputs,
            "scenario": r.scenario,
            **r.outputs,
            "issues": "; ".join(r.issues),
            "warnings": "; ".join(r.warnings),
            "status": "error" if r.issues else ("warning" if r.warnings else "ok"),
        }
        rows.append(row)
    return pd.DataFrame(rows)

def add_summary_sheets(df: pd.DataFrame, path: Path) -> None:
    wb = load_workbook(path)

    # Fehlerübersicht
    issues_expanded = []
    for _, row in df.iterrows():
        for issue in filter(None, row["issues"].split("; ")):
            issues_expanded.append(
                {
                    "issue": issue,
                    "scenario": row["scenario"],
                    "houseType": row["houseType"],
                    "roofArea": row["roofArea"],
                }
            )
    issue_df = pd.DataFrame(issues_expanded)
    summary_sheet = wb.create_sheet("Fehlerübersicht")
    if not issue_df.empty:
        grouped = issue_df.groupby("issue").agg(count=("issue", "size"))
        grouped.reset_index(inplace=True)
        for r_idx, row in enumerate([grouped.columns.tolist(), *grouped.values.tolist()], start=1):
            for c_idx, val in enumerate(row, start=1):
                summary_sheet.cell(row=r_idx, column=c_idx, value=val)

    # Handlungsempfehlungen
    rec_sheet = wb.create_sheet("Handlungsempfehlungen")
    recommendations = [
        "Speicherlogik erzeugt zu niedrigen Netzbezug – Lade-/Entladeverluste prüfen",
        "PV-Dimensionierung überschreitet Dachfläche – Begrenzung fehlt",
        "CO2-Ergebnis bei Wallbox ohne Einsparung – Verbrenner-Ersatz prüfen",
        "Break-even außerhalb Zielkorridor – Einspeisevergütung/EV-Einsparung prüfen",
    ]
    for idx, rec in enumerate(recommendations, start=1):
        rec_sheet.cell(row=idx, column=1, value=rec)

    wb.save(path)

def color_rows(path: Path, df: pd.DataFrame) -> None:
    wb = load_workbook(path)
    ws = wb["Testmatrix"]
    status_col = list(df.columns).index("status") + 1
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    for idx, status in enumerate(df["status"], start=2):  # header at row 1
        fill = green
        if status == "error":
            fill = red
        elif status == "warning":
            fill = yellow
        for col in range(1, len(df.columns) + 1):
            ws.cell(row=idx, column=col).fill = fill
    wb.save(path)

def main() -> None:
    data = load_data()
    inputs = build_inputs_matrix()
    all_results: List[TestResult] = []
    for inp in inputs:
        all_results.extend(scenario_calculations(data, inp))

    df = to_dataframe(all_results)
    
    # Erstelle test-Verzeichnis falls nicht vorhanden
    test_dir = Path(__file__).resolve().parent / "test"
    test_dir.mkdir(exist_ok=True)
    
    output_path = test_dir / "modernisierung_tests.xlsx"
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Testmatrix", index=False)
    color_rows(output_path, df)
    add_summary_sheets(df, output_path)
    print(f"Testmatrix geschrieben: {output_path.resolve()} ({len(df)} Zeilen)")

if __name__ == "__main__":
    main()
